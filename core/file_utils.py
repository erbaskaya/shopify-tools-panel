import csv, json, re, tempfile, uuid
from pathlib import Path
from fastapi import UploadFile

TMP_UPLOAD_DIR = Path(tempfile.gettempdir()) / 'shopify_tools_uploads'
TMP_EXPORT_DIR = Path(tempfile.gettempdir()) / 'shopify_tools_exports'
TMP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
TMP_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

def safe_filename(filename):
    filename = (filename or 'upload').replace('\\', '/').split('/')[-1]
    filename = re.sub(r'[^a-zA-Z0-9._-]+', '_', filename)
    return filename[:120] or 'upload'

def save_upload_file(upload_file: UploadFile, prefix='upload'):
    name = safe_filename(upload_file.filename)
    suffix = Path(name).suffix.lower()
    if suffix not in ('.csv', '.xlsx', '.xlsm', '.json'):
        raise Exception('Sadece CSV, XLSX, XLSM veya JSON dosyası yüklenebilir.')
    target = TMP_UPLOAD_DIR / f'{prefix}_{uuid.uuid4().hex[:10]}_{name}'
    upload_file.file.seek(0)
    target.write_bytes(upload_file.file.read())
    return target

def normalize_key(key):
    text = str(key or '').replace('\ufeff', '').strip().lower()
    for a,b in {'ı':'i','İ':'i','ğ':'g','ü':'u','ş':'s','ö':'o','ç':'c'}.items():
        text = text.replace(a,b)
    text = re.sub(r'[\s\-/]+', '_', text)
    text = re.sub(r'[^a-z0-9_]+', '', text)
    return text.strip('_')

def cell_to_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def normalize_row(headers, values):
    row = {}
    for i, header in enumerate(headers):
        key = normalize_key(header)
        if key:
            row[key] = cell_to_str(values[i] if i < len(values) else '')
    return row

def read_csv_rows(path):
    raw = Path(path).read_bytes()
    text = None
    for enc in ('utf-8-sig','utf-8','cp1254','latin-1'):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            pass
    if text is None:
        raise Exception('CSV okunamadı.')
    sample = text[:4096]
    delimiter = ';'
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=';,|\t').delimiter
    except Exception:
        first = text.splitlines()[0] if text.splitlines() else ''
        if first.count(',') > first.count(';'):
            delimiter = ','
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    if not rows:
        return []
    headers = rows[0]
    return [normalize_row(headers, r) for r in rows[1:] if any(cell_to_str(v) for v in r)]

def read_xlsx_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = next(it, None)
    if not headers:
        return []
    return [normalize_row(headers, r) for r in it if r and any(cell_to_str(v) for v in r)]

def read_rows(path):
    suffix = Path(path).suffix.lower()
    if suffix == '.csv': return read_csv_rows(path)
    if suffix in ('.xlsx','.xlsm'): return read_xlsx_rows(path)
    raise Exception('Bu işlem CSV, XLSX veya XLSM dosyası ister.')

def read_json(path):
    return json.loads(Path(path).read_text(encoding='utf-8'))

def first_value(row, *aliases, default=''):
    for alias in aliases:
        key = normalize_key(alias)
        if key in row and row[key] not in ('', None):
            return str(row[key]).strip()
    return default

def parse_bool(value, default=False):
    t = str(value or '').strip().lower()
    if t in ('1','true','yes','evet','ja','y','active','aktif','published'): return True
    if t in ('0','false','no','hayir','hayır','nein','n','draft','pasif','unpublished'): return False
    return default

def parse_int(value, default=None):
    t = str(value or '').strip()
    if not t: return default
    try: return int(float(t.replace(',','.')))
    except Exception: return default

def price_text(value):
    t = str(value or '').strip().replace('€','').replace(' ','').replace(',','.')
    return t

def split_values(value):
    t = str(value or '').strip()
    if not t: return []
    return [p.strip() for p in re.split(r'[|,]+', t) if p.strip()]

def make_export_path(filename):
    return TMP_EXPORT_DIR / safe_filename(filename)
